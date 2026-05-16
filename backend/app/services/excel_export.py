"""
Excel export service using openpyxl.
Produces a formatted multi-sheet workbook:
  Sheet 1: P&L Summary
  Sheet 2: Transaction Ledger
  Sheet 3: GST Summary
"""
import io
from typing import List
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# Brand colours
DARK_BG   = "1A1A1A"
GREEN     = "34D399"
RED       = "F87171"
YELLOW    = "FACC15"
HEADER_BG = "18181B"
ROW_ALT   = "27272A"
WHITE     = "FFFFFF"
GREY      = "A1A1AA"


def _border():
    thin = Side(style="thin", color="3F3F46")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_style(ws, row: int, cols: int, title: str, bg: str = HEADER_BG):
    ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
    cell = ws.cell(row=row, column=1, value=title)
    cell.font      = Font(bold=True, color=WHITE, size=13)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24


def _write_col_headers(ws, row: int, headers: list[str], bg: str = "27272A"):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font      = Font(bold=True, color=GREY, size=10)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center")
        cell.border    = _border()


def _inr(ws, row: int, col: int, value: float, positive_green: bool = False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = '₹#,##0.00'
    cell.alignment     = Alignment(horizontal="right")
    cell.border        = _border()
    if positive_green:
        cell.font = Font(color=GREEN if value >= 0 else RED, bold=True)
    return cell


def generate_excel(
    org_name: str,
    period_label: str,
    total_income: float,
    total_expenses: float,
    net_cashflow: float,
    category_totals: list,   # List[CategoryTotal]
    transactions: list,       # List[Transaction]
    gst_lines: list,          # List[GSTLine]
) -> bytes:
    wb = Workbook()

    # ── Sheet 1: P&L ─────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "P&L Summary"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 36
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 12

    row = 1
    _header_style(ws1, row, 3, f"Profit & Loss — {org_name}", DARK_BG); row += 1
    ws1.cell(row=row, column=1, value=f"Period: {period_label}").font = Font(color=GREY, italic=True); row += 2

    # Revenue
    ws1.cell(row=row, column=1, value="REVENUE").font = Font(bold=True, color=GREEN, size=11); row += 1
    income_cats = [c for c in category_totals if c.total > 0]
    for c in income_cats:
        ws1.cell(row=row, column=1, value=c.category).border = _border()
        ws1.cell(row=row, column=1).fill = PatternFill("solid", fgColor="18181B")
        _inr(ws1, row, 2, c.total)
        ws1.cell(row=row, column=3, value=f"{c.count} txns").font = Font(color=GREY, size=9)
        row += 1
    _inr(ws1, row, 2, total_income)
    ws1.cell(row=row, column=1, value="Total Revenue").font = Font(bold=True, color=WHITE)
    ws1.cell(row=row, column=1).fill = PatternFill("solid", fgColor="18181B"); row += 2

    # Expenses
    ws1.cell(row=row, column=1, value="EXPENSES").font = Font(bold=True, color=RED, size=11); row += 1
    expense_cats = [c for c in category_totals if c.total < 0]
    for c in expense_cats:
        ws1.cell(row=row, column=1, value=c.category).border = _border()
        ws1.cell(row=row, column=1).fill = PatternFill("solid", fgColor="18181B")
        _inr(ws1, row, 2, c.total)
        ws1.cell(row=row, column=3, value=f"{c.percentage:.1f}%").font = Font(color=GREY, size=9)
        row += 1
    _inr(ws1, row, 2, total_expenses)
    ws1.cell(row=row, column=1, value="Total Expenses").font = Font(bold=True, color=WHITE)
    ws1.cell(row=row, column=1).fill = PatternFill("solid", fgColor="18181B"); row += 2

    # Net
    ws1.cell(row=row, column=1, value="NET CASHFLOW").font = Font(bold=True, color=WHITE, size=12)
    ws1.cell(row=row, column=1).fill = PatternFill("solid", fgColor="18181B")
    _inr(ws1, row, 2, net_cashflow, positive_green=True)
    if total_income:
        ws1.cell(row=row, column=3, value=f"{net_cashflow/total_income*100:.1f}% margin").font = Font(color=GREY, size=9)
    row += 1

    # ── Sheet 2: Ledger ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Transaction Ledger")
    ws2.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [14, 42, 26, 14, 12, 12]):
        ws2.column_dimensions[get_column_letter(col.encode()[0] - 64)].width = w

    row = 1
    _header_style(ws2, row, 6, f"Transaction Ledger — {org_name}"); row += 1
    _write_col_headers(ws2, row, ["Date", "Description", "Category", "Amount", "GST Est.", "Reconciled"]); row += 1

    for i, t in enumerate(transactions):
        bg = "18181B" if i % 2 == 0 else "1F1F23"
        fill = PatternFill("solid", fgColor=bg)
        for col in range(1, 7):
            ws2.cell(row=row, column=col).fill = fill
            ws2.cell(row=row, column=col).border = _border()
        ws2.cell(row=row, column=1, value=str(t.date or ""))
        ws2.cell(row=row, column=2, value=t.description or "")
        ws2.cell(row=row, column=3, value=t.category or "")
        amt_cell = ws2.cell(row=row, column=4, value=t.amount)
        amt_cell.number_format = '₹#,##0.00'
        amt_cell.alignment     = Alignment(horizontal="right")
        amt_cell.font          = Font(color=GREEN if t.amount >= 0 else RED)
        gst_cell = ws2.cell(row=row, column=5, value=t.gst_amount or 0)
        gst_cell.number_format = '₹#,##0.00'
        gst_cell.alignment     = Alignment(horizontal="right")
        ws2.cell(row=row, column=6, value="✓" if t.is_reconciled else "—")
        ws2.cell(row=row, column=6).alignment = Alignment(horizontal="center")
        row += 1

    # ── Sheet 3: GST ─────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("GST Summary")
    ws3.sheet_view.showGridLines = False
    for col, w in zip(range(1, 7), [28, 16, 16, 14, 14, 14]):
        ws3.column_dimensions[get_column_letter(col)].width = w

    row = 1
    _header_style(ws3, row, 6, f"GST Input Tax Credit — {org_name}", "18181B"); row += 1
    _header_style(ws3, row, 6, f"Period: {period_label}", "1F1F23"); row += 1
    _write_col_headers(ws3, row, ["Category", "Taxable Amount", "GST (18%)", "CGST (9%)", "SGST (9%)", "IGST"]); row += 1

    for i, line in enumerate(gst_lines):
        bg = "18181B" if i % 2 == 0 else "1F1F23"
        fill = PatternFill("solid", fgColor=bg)
        for col in range(1, 7):
            ws3.cell(row=row, column=col).fill = fill
            ws3.cell(row=row, column=col).border = _border()
        ws3.cell(row=row, column=1, value=line.category)
        for col, val in zip(range(2, 7), [line.taxable_amount, line.gst_amount, line.cgst, line.sgst, line.igst]):
            c = ws3.cell(row=row, column=col, value=val)
            c.number_format = '₹#,##0.00'
            c.alignment     = Alignment(horizontal="right")
        row += 1

    # Totals row
    fill_total = PatternFill("solid", fgColor="27272A")
    ws3.cell(row=row, column=1, value="TOTAL").font = Font(bold=True, color=WHITE)
    ws3.cell(row=row, column=1).fill = fill_total
    for col, val in zip(range(2, 7), [
        sum(l.taxable_amount for l in gst_lines),
        sum(l.gst_amount for l in gst_lines),
        sum(l.cgst for l in gst_lines),
        sum(l.sgst for l in gst_lines),
        0.0,
    ]):
        c = ws3.cell(row=row, column=col, value=round(val, 2))
        c.number_format = '₹#,##0.00'
        c.alignment     = Alignment(horizontal="right")
        c.font          = Font(bold=True, color=YELLOW)
        c.fill          = fill_total

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
